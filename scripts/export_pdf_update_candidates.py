#!/usr/bin/env python3
"""
Create a review CSV of potential tree updates from supplied PDFs.

Output:
  reports/pdf-update-candidates.csv
"""

from __future__ import annotations

import csv
import gzip
import json
import re
from html import unescape
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
GRAPH_PATH = ROOT / "public" / "family-graph.json"
OUTPUT_PATH = ROOT / "reports" / "pdf-update-candidates.csv"
RAW_OUTPUT_PATH = ROOT / "reports" / "pdf-update-candidates.raw.csv"

# Explicit PDF_PATHS + EXPLICIT_EXTRA can exceed this; raise cap so MyHeritage DNA exports are not dropped after sort.
MAX_SOURCES = 80

PDF_PATHS = [
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/ee27502d-9f04-4c3a-b690-0456e8cd50a2/Hillman Tree Updated.pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/9578b2c7-76cf-4eef-bfff-daf91b14b970/AlpertsAndCohens31.pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/d34e5781-677b-4724-8481-db837d43f276/Geni - Relatives List.pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/77251b54-ae3c-4284-8107-6eb819525a18/111Geni - Relatives List.pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/e6b4305f-7423-4a7d-a5b8-f09a78d65d2f/:Geni - Relatives List.pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/0e1325d0-7fdd-49f4-a56d-188d1aa24eaf/11Geni - Relatives List.pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/7cdeb326-269f-4072-aeea-7ab254af976d/1Aalperovitch - Castrel.pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/b0eefe62-55f0-4ce6-80ac-af2e77d6d4d3/Tree Consistency Checker - Livnat Zaidman Web Site - MyHeritage.pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/be429099-bcac-48b9-ac2f-115b009a8dc0/Assif Hal Paternal Haplogroup Report - 23andMe.pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/25c074e2-531e-480c-9c9e-eef9986807db/1Assif Hal Paternal Haplogroup Report - 23andMe.pdf",
]

EXPLICIT_EXTRA_SOURCES = [
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/55df7f5e-7b13-465d-9182-5a7e3268e745/Ancestry Composition How It Works - 23andMe.pdf",
    "/Users/yaellivnatzaidman/Documents/LivnatGenealogyWorkspace/09-From-Downloads-Bundles/assif_hal_23andme_dna.rtfd",
    "/Users/yaellivnatzaidman/Documents/LivnatGenealogyWorkspace/01-Heritage-Project/Assif_23andMe_DNA_Matches_Complete.xlsx",
    "/Users/yaellivnatzaidman/Documents/LivnatGenealogyWorkspace/09-From-Downloads-Bundles/REORG_Livnat_Alperovich_Castro_2026-01-20/01_DNA_Data/Y-DNA/ Oded_YDNA_FTDNA.txt",
    "/Users/yaellivnatzaidman/Documents/LivnatGenealogyWorkspace/09-From-Downloads-Bundles/REORG_Livnat_Alperovich_Castro_2026-01-20/01_DNA_Data/Family_Finder/all-ftdna.numbers",
    "/Users/yaellivnatzaidman/Documents/LivnatGenealogyWorkspace/01-Heritage-Project/MyHeritage_raw_dna_data 2.csv",
    "/Users/yaellivnatzaidman/Documents/LivnatGenealogyWorkspace/01-Heritage-Project/assif_23andme_matches_REL_TO_ME_v1.xlsx",
    "/Users/yaellivnatzaidman/Documents/LivnatGenealogyWorkspace/01-Heritage-Project/assif_23andme_to_master_linkage_and_tidy_review_v1.xlsx",
    "/Users/yaellivnatzaidman/Documents/LivnatGenealogyWorkspace/01-Heritage-Project/Hal_DNA_23andMe_Analysis.xlsx",
    "/Users/yaellivnatzaidman/Documents/LivnatGenealogyWorkspace/01-Heritage-Project/37_Y_Zaidman_Chrom_Autoso_20260204.csv.gz",
    "/Users/yaellivnatzaidman/Documents/LivnatGenealogyWorkspace/01-Heritage-Project/IN131982_Chromosome_Browser_Results_20260204_ALL_plus_8_9_10.csv.gz",
    # MyHeritage site exports (Livnat Zaidman Web Site) — Chromosome Browser + DNA results PDFs
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/d9279bd5-e58d-4e2c-b31b-83da337e1c60/Chromosome Browser - Livnat Zaidman Web Site - MyHeritage.pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/efa35993-f6e4-4838-a27c-ce4be3f16f91/Chromosome Browser - Livnat Zaidman Web Site - MyHeritage (1).pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/795b6263-ad5f-48cb-8167-a47f71da2b55/Chromosome Browser - Livnat Zaidman Web Site - MyHeritage (2).pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/fd7ace7e-66a6-4bdf-a1e6-4e3b7d7c0d8c/Chromosome Browser - Livnat Zaidman Web Site - MyHeritage (3).pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/879f3b8f-373d-4fd4-b79e-3cac1788f980/Chromosome Browser - Livnat Zaidman Web Site - MyHeritage (4).pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/dc7dd43c-caf5-4681-84ce-236491380b67/DNA results - Livnat Zaidman Web Site - MyHeritage.pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/d5e30d4a-a954-42fc-967b-e8aa605bafb6/DNA results - Livnat Zaidman Web Site - MyHeritage (1).pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/752c8f2c-0a8e-4cd6-a15d-7f262176270f/DNA results - Livnat Zaidman Web Site - MyHeritage (2).pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/552e8bf6-0a70-4875-9aba-9af96d0cc04b/DNA results - Livnat Zaidman Web Site - MyHeritage (3).pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/91ea3650-5e62-48f7-b312-69aea2125ea7/DNA results - Livnat Zaidman Web Site - MyHeritage (4).pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/7992ff97-b427-4b35-8b13-ab3a0a2ce18c/DNA results - Livnat Zaidman Web Site - MyHeritage (5).pdf",
    "/Users/yaellivnatzaidman/Library/Application Support/Cursor/User/workspaceStorage/dc39c8d62d6cca1afa69d14726015523/pdfs/de58fe01-43ad-4780-9009-2323d871e90b/DNA results - Livnat Zaidman Web Site - MyHeritage (6).pdf",
    # Ben Zeadman — MyHeritage AutoClusters (MH-KN8886), March 2026
    "/Users/yaellivnatzaidman/Downloads/Ben Zeadman AutoClusters - MH-KN8886 - March 24 2026/Ben Zeadman AutoClusters - MH-KN8886 - March 24 2026.csv",
    "/Users/yaellivnatzaidman/Downloads/Ben Zeadman AutoClusters - MH-KN8886 - March 24 2026/Ben Zeadman AutoClusters - MH-KN8886 - March 24 2026.html",
    "/Users/yaellivnatzaidman/Downloads/Ben Zeadman AutoClusters - MH-KN8886 - March 24 2026/ReadMe.pdf",
]

EXTRA_SCAN_DIRS = [
    "/Users/yaellivnatzaidman/Documents/LivnatGenealogyWorkspace/01-Heritage-Project",
]

ALLOWED_EXTENSIONS = {".pdf", ".txt", ".csv", ".xlsx", ".rtfd", ".gz", ".html"}
SOURCE_KEYWORDS = (
    "dna",
    "match",
    "heritage",
    "autocluster",
    "zeadman",
    "ged",
    "family",
    "tree",
    "geni",
    "alper",
    "dubershtein",
    "duberstein",
    "ginzburg",
    "castro",
    "kastr",
    "zaidman",
    "livnat",
    "chromosome",
    "ftdna",
    "23andme",
)


CATEGORY_RULES: list[tuple[str, re.Pattern[str], str]] = [
    (
        "DNA",
        re.compile(
            r"\b(dna|haplogroup|ftdna|23andme|myheritage|match list|matching|mtdna|y-dna|autosomal)\b",
            re.IGNORECASE,
        ),
        "Validate person-level DNA evidence and add/remove Verified DNA links.",
    ),
    (
        "Lineage/Rabbi",
        re.compile(
            r"\b(rabbi|r'|av beit din|abd|maharam|rashi|gaon|ha-kohen|cohen)\b",
            re.IGNORECASE,
        ),
        "Review for Rabbi/Lineage/Famous tagging and title enrichment.",
    ),
    (
        "Surname/nee",
        re.compile(
            r"\b(nee|born|formerly|was|called|לבית|לשעבר)\b",
            re.IGNORECASE,
        ),
        "Review surname history (former/current/married) and manual field overrides.",
    ),
    (
        "Marriage-network",
        re.compile(
            r"\b(marriage|married|inter-family|interfamily|cousin marriage|uncle[- ]niece|first cousins|second cousins)\b",
            re.IGNORECASE,
        ),
        "Review kinship/network notes and in-family marriage signals.",
    ),
]


NAME_RE = re.compile(r"\b([A-Z][A-Za-z'\"().-]+(?:\s+[A-Z][A-Za-z'\"().-]+){1,5})\b")
BAD_NAME_WORDS = {
    "report",
    "haplogroup",
    "middle",
    "east",
    "central",
    "asia",
    "inheritance",
    "genomics",
    "http",
    "https",
    "print",
    "page",
    "paternal",
    "chromosome",
    "copyright",
    "father",
    "fathers",
    "cousin",
    "removed",
    "sheet",
    "rank",
    "relationship",
    "segments",
    "side",
    "notes",
    "analysis",
    "match",
    "matches",
    "tree",
}


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def normalize_name(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (text or "").lower()).strip()


@dataclass
class Candidate:
    source_file: str
    category: str
    candidate_name: str
    matched_person_ids: str
    confidence: str
    evidence_snippet: str
    suggested_action: str


def read_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    chunks: list[str] = []
    for page in reader.pages:
        try:
            page_text = page.extract_text() or ""
        except Exception:
            page_text = ""
        if page_text.strip():
            chunks.append(page_text)
    return "\n".join(chunks)


def read_txt_text(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return ""


def read_csv_text(path: Path, max_rows: int = 4000) -> str:
    lines: list[str] = []
    try:
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= max_rows:
                    break
                row_text = " | ".join(str(cell) for cell in row if str(cell).strip())
                if row_text:
                    lines.append(row_text)
    except Exception:
        return ""
    return "\n".join(lines)


def read_html_text(path: Path, max_chars: int = 900_000) -> str:
    """Strip tags for name/DNA keyword matching (MyHeritage AutoClusters exports, etc.)."""
    try:
        raw = path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return ""
    if len(raw) > max_chars:
        raw = raw[:max_chars]
    raw = re.sub(r"<script\b[^>]*>.*?</script>", " ", raw, flags=re.DOTALL | re.IGNORECASE)
    raw = re.sub(r"<style\b[^>]*>.*?</style>", " ", raw, flags=re.DOTALL | re.IGNORECASE)
    raw = re.sub(r"<[^>]+>", " ", raw)
    raw = unescape(raw)
    return normalize_space(raw)


def read_gz_csv_text(path: Path, max_rows: int = 4000) -> str:
    lines: list[str] = []
    try:
        with gzip.open(path, "rt", encoding="utf-8", errors="ignore", newline="") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= max_rows:
                    break
                row_text = " | ".join(str(cell) for cell in row if str(cell).strip())
                if row_text:
                    lines.append(row_text)
    except Exception:
        return ""
    return "\n".join(lines)


def read_xlsx_text(path: Path, max_rows_per_sheet: int = 2500, max_sheets: int = 6) -> str:
    lines: list[str] = []
    try:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
    except Exception:
        return ""

    try:
        for sheet_idx, sheet in enumerate(workbook.worksheets):
            if sheet_idx >= max_sheets:
                break
            lines.append(f"Sheet: {sheet.title}")
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= max_rows_per_sheet:
                    break
                row_text = " | ".join(str(cell) for cell in row if cell is not None and str(cell).strip())
                if row_text:
                    lines.append(row_text)
    finally:
        workbook.close()
    return "\n".join(lines)


def strip_rtf_markup(text: str) -> str:
    # Keep a lightweight cleanup; this is sufficient for candidate extraction.
    text = re.sub(r"\\'[0-9a-fA-F]{2}", " ", text)
    text = re.sub(r"\\[a-zA-Z]+\d* ?", " ", text)
    text = re.sub(r"[{}]", " ", text)
    return normalize_space(text)


def read_rtfd_text(path: Path) -> str:
    rtf_file = path / "TXT.rtf"
    if not rtf_file.exists():
        return ""
    raw = read_txt_text(rtf_file)
    return strip_rtf_markup(raw)


def read_source_text(path: Path) -> str:
    lower_name = path.name.lower()
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return read_pdf_text(path)
    if suffix == ".txt":
        return read_txt_text(path)
    if suffix == ".csv":
        return read_csv_text(path)
    if suffix == ".html":
        return read_html_text(path)
    if suffix == ".xlsx":
        return read_xlsx_text(path)
    if suffix == ".rtfd":
        return read_rtfd_text(path)
    if suffix == ".gz" and lower_name.endswith(".csv.gz"):
        return read_gz_csv_text(path)
    return ""


def collect_sources() -> tuple[list[Path], list[str]]:
    sources: list[Path] = []
    seen: set[str] = set()
    skipped: list[str] = []
    explicit = {str(Path(p)) for p in (PDF_PATHS + EXPLICIT_EXTRA_SOURCES)}

    def add_path(p: Path) -> None:
        key = str(p)
        if key in seen:
            return
        seen.add(key)
        if p.suffix.lower() == ".numbers":
            skipped.append(f"{p.name} (unsupported .numbers)")
            return
        if p.suffix.lower() not in ALLOWED_EXTENSIONS and not str(p).lower().endswith(".csv.gz"):
            return
        if str(p) not in explicit:
            lower_name = p.name.lower()
            if not any(k in lower_name for k in SOURCE_KEYWORDS):
                return
        sources.append(p)

    for raw_path in PDF_PATHS + EXPLICIT_EXTRA_SOURCES:
        p = Path(raw_path)
        if p.exists():
            add_path(p)
        else:
            skipped.append(f"{p.name} (missing)")

    for raw_dir in EXTRA_SCAN_DIRS:
        d = Path(raw_dir)
        if not d.exists():
            skipped.append(f"{d.name} (missing dir)")
            continue
        for child in d.rglob("*"):
            if len(sources) >= MAX_SOURCES:
                break
            if child.is_file():
                add_path(child)
            elif child.is_dir() and child.suffix.lower() == ".rtfd":
                add_path(child)

    # Prevent very large scans from stalling; keep the most relevant files by name.
    sources.sort(key=lambda p: p.name.lower())
    if len(sources) > MAX_SOURCES:
        skipped.append(f"{len(sources) - MAX_SOURCES} sources omitted by cap")
        sources = sources[:MAX_SOURCES]

    return sources, skipped


def load_people_index() -> list[tuple[str, str, str]]:
    if not GRAPH_PATH.exists():
        return []
    data = json.loads(GRAPH_PATH.read_text(encoding="utf-8"))
    rows: list[tuple[str, str, str]] = []
    for person in data.get("persons", []):
        full_name = person.get("fullName", "")
        if not full_name:
            continue
        normalized = normalize_name(full_name)
        if len(normalized) < 4:
            continue
        if len([t for t in normalized.split(" ") if t]) < 2:
            continue
        rows.append((person.get("id", ""), full_name, normalized))
    return rows


def best_person_matches(candidate_name: str, people_index: list[tuple[str, str, str]]) -> str:
    norm = normalize_name(candidate_name)
    if not norm or len(norm) < 5:
        return ""
    tokens = [t for t in norm.split(" ") if t]
    if not tokens:
        return ""

    matches: list[tuple[str, str]] = []
    for pid, full_name, pnorm in people_index:
        if len(pnorm) < 4 or len(norm) < 4:
            continue
        if norm == pnorm or (len(norm) >= 6 and norm in pnorm) or (len(pnorm) >= 6 and pnorm in norm):
            matches.append((pid, full_name))
            continue
        # token overlap fallback
        overlap = sum(1 for token in tokens if token in pnorm.split(" "))
        if overlap >= max(2, len(tokens) - 1):
            matches.append((pid, full_name))

    if not matches:
        return ""
    return " | ".join(f"{pid}:{name}" for pid, name in matches[:3])


def infer_confidence(snippet: str, matched_ids: str) -> str:
    s = snippet.lower()
    if matched_ids and (
        "myheritage" in s
        or "ftdna" in s
        or "relatives list" in s
        or "geni" in s
        or "match" in s
    ):
        return "high"
    if matched_ids:
        return "medium"
    return "low"


def looks_like_person_name(candidate_name: str) -> bool:
    name = normalize_space(candidate_name)
    tokens = [t for t in re.split(r"\s+", name) if t]
    if len(tokens) < 2:
        return False
    lowered = [re.sub(r"[^a-z]+", "", t.lower()) for t in tokens]
    if any(len(t) <= 1 for t in lowered if t):
        return False
    if any(t in BAD_NAME_WORDS for t in lowered if t):
        return False
    if "http" in name.lower() or "www" in name.lower():
        return False
    return True


def extract_candidates_from_text(
    source_name: str,
    text: str,
    people_index: list[tuple[str, str, str]],
) -> list[Candidate]:
    results: list[Candidate] = []
    seen: set[tuple[str, str, str]] = set()
    lines = [normalize_space(line) for line in text.splitlines()]
    lines = [line for line in lines if line]
    source_lower = source_name.lower()
    dna_only_source = any(
        token in source_lower
        for token in (
            "23andme",
            "dna",
            "ftdna",
            "chrom",
            "myheritage_raw_dna_data",
            "autocluster",
        )
    )

    for i, line in enumerate(lines):
        context = " ".join(lines[max(0, i - 1) : min(len(lines), i + 2)])
        context = normalize_space(context)
        if len(context) < 20:
            continue

        for category, pattern, action in CATEGORY_RULES:
            if dna_only_source and category != "DNA":
                continue
            if not pattern.search(context):
                continue

            names = NAME_RE.findall(context)
            if not names:
                continue

            for name in names:
                cleaned_name = normalize_space(name)
                if len(cleaned_name) < 6 or not looks_like_person_name(cleaned_name):
                    continue
                matched_ids = best_person_matches(cleaned_name, people_index)
                if not matched_ids:
                    continue
                key = (source_name, category, cleaned_name.lower())
                if key in seen:
                    continue
                seen.add(key)
                results.append(
                    Candidate(
                        source_file=source_name,
                        category=category,
                        candidate_name=cleaned_name,
                        matched_person_ids=matched_ids,
                        confidence=infer_confidence(context, matched_ids),
                        evidence_snippet=context[:220],
                        suggested_action=action,
                    )
                )
    return results


def write_csv(path: Path, rows: Iterable[Candidate]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "source_file",
                "category",
                "candidate_name",
                "matched_person_ids",
                "confidence",
                "evidence_snippet",
                "suggested_action",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.source_file,
                    row.category,
                    row.candidate_name,
                    row.matched_person_ids,
                    row.confidence,
                    row.evidence_snippet,
                    row.suggested_action,
                ]
            )


def main() -> None:
    people_index = load_people_index()
    all_rows: list[Candidate] = []
    scanned: list[str] = []
    sources, skipped = collect_sources()
    for path in sources:
        scanned.append(path.name)
        text = read_source_text(path)
        if not text.strip():
            continue
        if len(text) > 300000:
            text = text[:300000]
        all_rows.extend(extract_candidates_from_text(path.name, text, people_index))

    # Sort high confidence first, then by category/name.
    rank = {"high": 0, "medium": 1, "low": 2}
    all_rows.sort(key=lambda r: (rank.get(r.confidence, 3), r.category, r.candidate_name))
    filtered_rows = [row for row in all_rows if row.confidence in {"high", "medium"}]
    write_csv(RAW_OUTPUT_PATH, all_rows)
    write_csv(OUTPUT_PATH, filtered_rows)
    print(f"Scanned sources ({len(scanned)}): {', '.join(scanned)}")
    if skipped:
        print(f"Skipped/missing ({len(skipped)}): {', '.join(skipped)}")
    print(f"Wrote {len(filtered_rows)} filtered candidate rows -> {OUTPUT_PATH}")
    print(f"Wrote {len(all_rows)} raw candidate rows -> {RAW_OUTPUT_PATH}")


if __name__ == "__main__":
    main()

